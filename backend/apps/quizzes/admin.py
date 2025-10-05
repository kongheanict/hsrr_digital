import pandas as pd
from django.contrib import admin, messages
from django import forms
from django.shortcuts import render, redirect
from django.http import HttpResponse, FileResponse
from django.urls import path, reverse
from django.utils.translation import gettext_lazy as trans
from django.contrib.admin.models import LogEntry, CHANGE
from django.contrib.contenttypes.models import ContentType
import nested_admin
from django_ckeditor_5.widgets import CKEditor5Widget
from datetime import timedelta, datetime
from django.utils.dateparse import parse_datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.db import transaction
from django.utils import timezone
import difflib
import re

from .models import QuizCategory, Quiz, Question, AnswerOption, StudentResponse, QuizAttempt, QuizAttemptQuestion
from apps.teachers.models import Teacher
from apps.classes.models import SchoolClass
from .forms import ExcelImportForm
from apps.core.models import AcademicYear
from apps.classes.models import HomeroomTeacher

from django.template.response import TemplateResponse

# -------------------------
# Small Textarea Widget
# -------------------------
class SmallTextarea(forms.Textarea):
    def __init__ (self, *args, **kwargs):
        kwargs.setdefault("attrs", {}).update({"rows": 2, "style": "width: 90%;"})
        super().__init__ (*args, **kwargs)

# -------------------------
# Forms
# -------------------------
class QuizForm(forms.ModelForm):
    start_time = forms.DateTimeField(
        required=False,
        widget=forms.DateTimeInput(
            attrs={"type": "datetime-local", "step": "1"},
            format="%Y-%m-%dT%H:%M",
        ),
        input_formats=["%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S"],
    )

    class Meta:
        model = Quiz
        fields = "__all__"
        widgets = {
            "description": SmallTextarea(),
            "time_limit": forms.TextInput(
                attrs={"placeholder": "e.g., 00:30:00 for 30 minutes"}
            ),
        }

    def __init__ (self, *args, **kwargs):
        self.request = kwargs.pop("request", None)
        super().__init__ (*args, **kwargs)
        if self.request and not self.request.user.is_superuser and "teacher" in self.fields:
            self.fields["teacher"].disabled = True
            if not self.instance.pk:
                teacher = Teacher.objects.filter(user=self.request.user).first()
                if teacher:
                    self.initial["teacher"] = teacher

    def clean_start_time(self):
        start_time = self.cleaned_data.get("start_time")
        if start_time and isinstance(start_time, (list, tuple)):
            raise forms.ValidationError( trans ("Start time must be a single datetime, not a list."))
        return start_time

    def clean_time_limit(self):
        time_limit = self.cleaned_data.get("time_limit")
        if time_limit:
            try:
                h, m, s = map(int, str(time_limit).split(":"))
                if h < 0 or m < 0 or s < 0 or m > 59 or s > 59:
                    raise ValueError
                return time_limit
            except (ValueError, AttributeError):
                raise forms.ValidationError( trans ("Invalid time limit format. Use HH:MM:SS (e.g., 00:30:00)."))
        return time_limit

    def clean(self):
        cleaned_data = super().clean()
        if self.request:
            if not self.request.user.is_superuser:
                teacher = Teacher.objects.filter(user=self.request.user).first()
                if not teacher:
                    raise forms.ValidationError( trans ("No teacher profile associated with this user."))
                cleaned_data["teacher"] = teacher
            else:
                if not cleaned_data.get("teacher"):
                    raise forms.ValidationError( trans ("Superadmin must assign a teacher to this quiz."))
        return cleaned_data

class QuestionForm(forms.ModelForm):
    class Meta:
        model = Question
        fields = "__all__"
        widgets = {"text": CKEditor5Widget()}

class AnswerOptionForm(forms.ModelForm):
    class Meta:
        model = AnswerOption
        fields = "__all__"
        widgets = {"text": SmallTextarea()}

# -------------------------
# Nested Inlines
# -------------------------
class AnswerOptionInline(nested_admin.NestedTabularInline):
    model = AnswerOption
    form = AnswerOptionForm
    extra = 0
    min_num = 1
    classes = ("collapse",)

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)

        class AnswerOptionFormset(formset):
            def clean(self):
                super().clean()
                if not self.is_valid():
                    return
                question_type = self.instance.question.question_type if self.instance and hasattr(self.instance, 'question') and self.instance.question else None
                if question_type in ["MCQ_SINGLE", "MCQ_MULTI"]:
                    correct_options = sum(
                        1 for form in self.forms if form.cleaned_data.get("is_correct", False)
                    )
                    if correct_options < 1:
                        raise forms.ValidationError(
                             trans ("At least one option must be marked as correct for this question.")
                        )
                    if question_type == "MCQ_SINGLE" and correct_options > 1:
                        raise forms.ValidationError(
                             trans ("Only one option can be marked as correct for single-choice questions.")
                        )

        return AnswerOptionFormset

class QuestionInline(nested_admin.NestedTabularInline):
    model = Question
    form = QuestionForm
    inlines = [AnswerOptionInline]
    extra = 0
    show_change_link = True

# -------------------------
# Admin Classes
# -------------------------
@admin.register(QuizCategory)
class QuizCategoryAdmin(admin.ModelAdmin):
    list_display = ("name", "description")
    search_fields = ("name",)

@admin.register(Quiz)
class QuizAdmin(nested_admin.NestedModelAdmin):
    form = QuizForm
    list_display = ("title", "category", "teacher", "status", "start_time", "time_limit", "created_at")
    search_fields = ("title",)
    list_filter = ("category", "classes", "status")
    inlines = [QuestionInline]
    readonly_fields = ("created_at",)
    autocomplete_fields = ("teacher", "category")
    filter_horizontal = ("classes",)
    change_list_template = "admin/quizzes/quiz_changelist.html"
    change_form_template = "admin/quizzes/quiz_change.html"  # Custom change form template
    actions = ['check_student_response', 'export_student_responses', 'recalculate_all_attempts']

    class Media:
        css = {"all": ("admin_custom.css",)}

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        teacher = Teacher.objects.filter(user=request.user).first()
        if teacher:
            return qs.filter(teacher=teacher)
        return qs.none()

    def get_form(self, request, obj=None, **kwargs):
        form_class = super().get_form(request, obj, **kwargs)

        class FormWithRequest(form_class):
            def __init__(self2, *args, **kwargs2):
                self2.request = request   # attach request here
                super().__init__(*args, **kwargs2)

        return FormWithRequest

    def save_model(self, request, obj, form, change):
        if not request.user.is_superuser:
            teacher = Teacher.objects.filter(user=request.user).first()
            if teacher:
                obj.teacher = teacher
        super().save_model(request, obj, form, change)

    def check_student_response(self, request, queryset):
        if queryset.count() > 1:
            self.message_user(
                request, 
                trans ("Please select exactly one quiz to view student responses."),
                level=messages.WARNING,
            )
            return
        # Check permission
        if not request.user.has_perm('quizzes.add_quiz'):
            self.message_user(
                request,
                trans ("You do not have permission to perform this action."),
                level=messages.ERROR,
            )
            return

        # Collect data for rendering
        quiz_data = []
        for quiz in queryset:
            # Get all questions for the quiz, ordered by 'order'
            questions = quiz.questions.all().order_by('order')
            question_headers = [f"Q{i+1}" for i in range(len(questions))]
            question_map = {q.id: f"Q{i+1}" for i, q in enumerate(questions)}

            # Get all attempts for the quiz
            attempts = quiz.attempts.select_related('student').prefetch_related('responses__question', 'responses__selected_options')
            # Group responses by student via attempts
            student_responses = {}
            for attempt in attempts:
                student = attempt.student
                student_name = f"{student.family_name} {student.given_name}"
                student_id = student.student_id
                if student_name not in student_responses:
                    student_responses[student_name] = {'student_id': student_id, 'responses': {}, 'attempt': attempt}
                # Get responses for this attempt
                responses = attempt.responses.select_related('question').prefetch_related('selected_options').order_by('question__order')
                for response in responses:
                    question_key = question_map.get(response.question.id, f"Q{response.question.order}")
                    response_data = {
                        'question_text': response.question.text,
                        'question_type': response.question.question_type,
                        'points_earned': response.points_earned,
                        'answer': response.text_answer or ', '.join(opt.text for opt in response.selected_options.all()) or trans ("No answer"),
                    }
                    student_responses[student_name]['responses'][question_key] = response_data

            # Get all quiz attempts for scores
            attempt_scores = {attempt.student.student_id: {
                'score': attempt.score,
                'completed_at': attempt.completed_at,
            } for attempt in attempts}

            quiz_data.append({
                'quiz_title': quiz.title,
                'question_headers': question_headers,
                'student_responses': student_responses,
                'student_attempts': attempt_scores,
            })

        # Render response page
        context = {
            'site_header' : "ប្រព័ន្ធគ្រប់គ្រងទិន្នន័យ",
            'title': trans ('បញ្ជីចម្លើយ និងពិន្ទុសិស្សបានធ្វើកម្រងសំណួរ'),
            'quiz_data': quiz_data,
            'opts': self.model._meta,
        }
        return TemplateResponse(
            request,
            'admin/quizzes/student_response.html',
            context,
        )

    check_student_response.short_description = "ត្រួតពិនិត្យពិន្ទុសិស្សដែលបានធ្វើកម្រងសំណួរ"

    def export_student_responses(self, request, queryset):
        """Export student responses and total scores for selected quizzes as Excel."""
        # Check permission
        if not request.user.has_perm('quizzes.add_quiz'):
            self.message_user(
                request,
                trans("You do not have permission to perform this action."),
                level=messages.ERROR,
            )
            return

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)  # Remove default sheet

        # Define current_year for use in both superuser and non-superuser paths
        try:
            current_year = AcademicYear.objects.get(status=True)
        except (AcademicYear.DoesNotExist, AcademicYear.MultipleObjectsReturned):
            current_year = AcademicYear.objects.order_by('-start_date').first()
            if not current_year:
                # If no academic year exists, set a warning and use empty queryset for non-superusers
                if not request.user.is_superuser:
                    self.message_user(
                        request,
                        trans("No active academic year found. Please configure an academic year."),
                        level=messages.WARNING,
                    )

        for quiz in queryset:
            # Create a new sheet for each quiz
            ws = wb.create_sheet(title=quiz.title[:31])  # Sheet title limited to 31 chars

            # Add report title
            ws.cell(row=1, column=1, value=str(trans("របាយការណ៍ពិន្ទុសិស្សសម្រាប់កម្រងសំណួរ: {}")).format(quiz.title))
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            title_cell = ws.cell(1, 1)
            title_cell.font = Font(bold=True, size=14, name="Khmer OS")
            title_cell.alignment = Alignment(horizontal="center")

            # Add date
            ws.cell(row=2, column=1, value=str(trans("កាលបរិច្ឆេទ: {}")).format(datetime.now().strftime("%d-%m-%Y")))
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
            date_cell = ws.cell(2, 1)
            date_cell.alignment = Alignment(horizontal="center")

            # Headers
            headers = [
                str(trans("អត្តលេខសិស្ស")),
                str(trans("ឈ្មោះសិស្ស")),
                str(trans("ភេទ")),
                str(trans("ថ្ងៃខែឆ្នាំកំណើត")),
                str(trans("ថ្នាក់")),
                str(trans("ពិន្ទុសរុប")),
            ]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num, value=header)
                cell.font = Font(bold=True, name="Khmer OS")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[cell.column_letter].width = 20  # Adjust column width

            # Get all quiz attempts for the quiz
            attempts = quiz.attempts.select_related('student')

            # Filter based on user permissions (e.g., homeroom teacher)
            if not request.user.is_superuser:
                try:
                    teacher = Teacher.objects.get(user=request.user)
                    if current_year:
                        homeroom_classes = HomeroomTeacher.objects.filter(
                            teacher=teacher,
                            academic_year=current_year
                        ).values_list('school_class__id', flat=True)
                        attempts = attempts.filter(
                            student__enrollments__school_class__id__in=homeroom_classes,
                            student__enrollments__academic_year=current_year,
                            student__enrollments__status='ACTIVE'
                        ).distinct()
                    else:
                        attempts = attempts.none()
                except Teacher.DoesNotExist:
                    attempts = attempts.none()

            row_num = 5
            for attempt in attempts:
                student = attempt.student
                enrollment = student.enrollments.filter(status='ACTIVE', academic_year=current_year).first() if current_year else None
                class_name = enrollment.school_class.name if enrollment else str(trans("មិនស្គាល់"))

                # Add row data
                row = [
                    student.student_id,
                    f"{student.family_name} {student.given_name}",
                    student.get_gender_display(),
                    student.date_of_birth.strftime("%d-%m-%Y") if student.date_of_birth else "",
                    class_name,
                    attempt.score,
                ]
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = Font(name="Khmer OS Siemreap")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                row_num += 1

            # Add borders to the table
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows(min_row=4, max_row=row_num-1, min_col=1, max_col=6):
                for cell in row:
                    cell.border = thin_border

            # Freeze header row
            ws.freeze_panes = ws['A5']

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=str(trans("ពិន្ទុសិស្ស.xlsx")))

    export_student_responses.short_description = str(trans("ទាញយកពិន្ទុសិស្សជា Excel"))

    def save_related(self, request, form, formsets, change):
        """Detect changes to AnswerOption.is_correct or SHORT question/answer text and recalculate scores."""
        quiz = form.instance
        original_answer_options = {}
        original_questions = {}
        if change:
            for question in quiz.questions.all():
                original_questions[question.id] = {
                    'text': question.text,
                    'question_type': question.question_type
                }
                for option in question.options.all():
                    original_answer_options[option.id] = {
                        'is_correct': option.is_correct,
                        'text': option.text
                    }

        super().save_related(request, form, formsets, change)

        answer_changed = False
        for question_formset in formsets:
            if hasattr(question_formset, 'forms'):
                for question_form in question_formset.forms:
                    question = question_form.instance
                    if isinstance(question, Question) and question.pk:
                        original = original_questions.get(question.id, {})
                        if question.question_type == 'SHORT' and original.get('text') != question.text:
                            answer_changed = True
                            LogEntry.objects.log_action(
                                user_id=request.user.id,
                                content_type_id=ContentType.objects.get_for_model(Question).pk,
                                object_id=question.pk,
                                object_repr=str(question),
                                action_flag=CHANGE,
                                change_message= trans ("Updated SHORT question text to: %s") % question.text
                            )
                    if hasattr(question_formset, 'nested_formsets'):
                        for answer_formset in question_formset.nested_formsets:
                            for answer_form in answer_formset.forms:
                                answer = answer_form.instance
                                if isinstance(answer, AnswerOption) and answer.pk:
                                    original_option = original_answer_options.get(answer.pk, {})
                                    if original_option.get('is_correct') != answer.is_correct:
                                        answer_changed = True
                                        LogEntry.objects.log_action(
                                            user_id=request.user.id,
                                            content_type_id=ContentType.objects.get_for_model(AnswerOption).pk,
                                            object_id=answer.pk,
                                            object_repr=str(answer),
                                            action_flag=CHANGE,
                                            change_message= trans ("Changed is_correct to %s for option: %s") % (
                                                answer.is_correct, answer.text
                                            )
                                        )
                                    if (hasattr(answer, 'question') and answer.question and 
                                        answer.question.question_type == 'SHORT' and 
                                        original_option.get('text') != answer.text):
                                        answer_changed = True
                                        LogEntry.objects.log_action(
                                            user_id=request.user.id,
                                            content_type_id=ContentType.objects.get_for_model(AnswerOption).pk,
                                            object_id=answer.pk,
                                            object_repr=str(answer),
                                            action_flag=CHANGE,
                                            change_message= trans ("Updated SHORT answer text to: %s") % answer.text
                                        )

        if answer_changed:
            attempts = QuizAttempt.objects.filter(quiz=quiz)
            for attempt in attempts:
                self.recalculate_quiz_attempt_score(attempt, request)
            if attempts:
                messages.info(request,  trans ("Recalculated scores for %d QuizAttempt(s) due to answer changes.") % len(attempts))

    def recalculate_quiz_attempt_score(self, attempt, request=None):
        """Recalculate the total score for a QuizAttempt based on StudentResponse."""
        quiz = attempt.quiz
        student = attempt.student
        # Updated to filter by attempt
        student_responses = StudentResponse.objects.filter(
            attempt=attempt
        ).select_related('question').prefetch_related('selected_options')
        
        total_score = 0
        # Loop over selected questions for the attempt
        selected_questions = attempt.selected_questions.select_related('question').order_by('order')
        for aq in selected_questions:
            question = aq.question
            response = student_responses.filter(question=question).first()
            score = 0
            if response:
                user_answer = list(response.selected_options.values_list('id', flat=True)) if response.selected_options.exists() else []
                user_text_answer = response.text_answer if response.text_answer else None

                if question.question_type in ["MCQ_SINGLE", "MCQ_MULTI"]:
                    correct_option_ids = [opt.id for opt in question.options.filter(is_correct=True)]
                    if question.question_type == "MCQ_SINGLE":
                        if len(user_answer) == 1 and user_answer[0] in correct_option_ids:
                            score = question.points
                    elif question.question_type == "MCQ_MULTI":
                        # Updated scoring logic to match views.py (partial credit with penalty)
                        all_options = set(question.options.values_list('id', flat=True))
                        correct_options = set(correct_option_ids)
                        wrong_options = all_options - correct_options
                        user_answers = set(user_answer)

                        num_correct = len(correct_options)
                        num_wrong = len(wrong_options)

                        num_correct_selected = len(user_answers & correct_options)
                        num_wrong_selected = len(user_answers & wrong_options)

                        score_from_correct = (question.points * num_correct_selected / num_correct) if num_correct else 0
                        penalty_from_wrong = (question.points * num_wrong_selected / num_wrong) if num_wrong else 0

                        score = score_from_correct - penalty_from_wrong
                        score = max(0, min(score, question.points))
                elif question.question_type == "SHORT" and user_text_answer:
                    # Updated to match views.py SHORT answer scoring
                    correct_options = question.options.filter(is_correct=True)
                    student_text = user_text_answer.strip().lower()

                    numeric_match = False
                    for opt in correct_options:
                        numbers = re.findall(r'\d+', student_text)
                        if str(opt.text.strip()) in numbers:
                            numeric_match = True
                            break

                    if numeric_match:
                        score = question.points
                    else:
                        keywords = []
                        for opt in correct_options:
                            keywords.extend(re.findall(r'\w+', opt.text.strip().lower()))
                        if keywords:
                            points_per_keyword = question.points / len(keywords)
                            match_count = sum(1 for kw in keywords if kw in student_text)
                            score = round(points_per_keyword * match_count, 2)
                        else:
                            if correct_options.exists():
                                similarity_scores = [
                                    difflib.SequenceMatcher(None, student_text, opt.text.strip().lower()).ratio()
                                    for opt in correct_options
                                ]
                                score = round(question.points * max(similarity_scores), 2)

                # Update response points_earned
                if response.points_earned != score:
                    response.points_earned = score
                    response.save(update_fields=['points_earned'])

            total_score += score

        if attempt.score != total_score:
            attempt.score = round(total_score, 2)
            attempt.save()
            if request:
                LogEntry.objects.log_action(
                    user_id=request.user.id,
                    content_type_id=ContentType.objects.get_for_model(attempt).pk,
                    object_id=attempt.pk,
                    object_repr=str(attempt),
                    action_flag=CHANGE,
                    change_message=str(trans("គណនាពិន្ទុឡើងវិញទៅ %s សម្រាប់កម្រងសំណួរ: %s") % (total_score, quiz.title))
                )
        return total_score

    def recalculate_single_quiz(self, request, object_id):
        """Recalculate scores for all QuizAttempts of a single quiz."""
        if not request.user.has_perm('quizzes.add_quiz'):
            self.message_user(
                request,
                trans("You do not have permission to perform this action."),
                level=messages.ERROR,
            )
            return redirect('admin:quizzes_quiz_changelist')

        try:
            quiz = self.get_object(request, object_id)
            if not quiz:
                self.message_user(
                    request,
                    trans("Quiz not found."),
                    level=messages.ERROR,
                )
                return redirect('admin:quizzes_quiz_changelist')

            try:
                current_year = AcademicYear.objects.get(status=True)
            except (AcademicYear.DoesNotExist, AcademicYear.MultipleObjectsReturned):
                current_year = AcademicYear.objects.order_by('-start_date').first()
                if not current_year:
                    self.message_user(
                        request,
                        trans("No active academic year found. Please configure an academic year."),
                        level=messages.WARNING,
                    )

            attempts = quiz.attempts.select_related('student')
            if not request.user.is_superuser:
                try:
                    teacher = Teacher.objects.get(user=request.user)
                    if current_year:
                        homeroom_classes = HomeroomTeacher.objects.filter(
                            teacher=teacher,
                            academic_year=current_year
                        ).values_list('school_class__id', flat=True)
                        attempts = attempts.filter(
                            student__enrollments__school_class__id__in=homeroom_classes,
                            student__enrollments__academic_year=current_year,
                            student__enrollments__status='ACTIVE'
                        ).distinct()
                    else:
                        attempts = attempts.none()
                except Teacher.DoesNotExist:
                    attempts = attempts.none()

            total_attempts = attempts.count()
            updated_attempts = 0
            for attempt in attempts:
                old_score = attempt.score
                new_score = self.recalculate_quiz_attempt_score(attempt, request=request)
                if old_score != new_score:
                    updated_attempts += 1

            if total_attempts > 0:
                self.message_user(
                    request,
                    str(trans("បានគណនាពិន្ទុឡើងវិញសម្រាប់ %d QuizAttempt(s) និងធ្វើបច្ចុប្បន្នភាព %d សម្រាប់កម្រងសំណួរ: %s") % (total_attempts, updated_attempts, quiz.title)),
                    level=messages.SUCCESS,
                )
            else:
                self.message_user(
                    request,
                    str(trans("គ្មាន QuizAttempt សម្រាប់គណនាឡើងវិញសម្រាប់កម្រងសំណួរ: %s") % quiz.title),
                    level=messages.WARNING,
                )

            return redirect('admin:quizzes_quiz_change', object_id)

        except Exception as e:
            self.message_user(
                request,
                str(trans("កំហុសក្នុងការគណនាពិន្ទុឡើងវិញ: %s") % str(e)),
                level=messages.ERROR,
            )
            return redirect('admin:quizzes_quiz_changelist')

    def recalculate_all_attempts(self, request, queryset):
        """Admin action to recalculate scores for all QuizAttempts of selected quizzes."""
        if not request.user.has_perm('quizzes.add_quiz'):
            self.message_user(
                request,
                trans("You do not have permission to perform this action."),
                level=messages.ERROR,
            )
            return

        try:
            current_year = AcademicYear.objects.get(status=True)
        except (AcademicYear.DoesNotExist, AcademicYear.MultipleObjectsReturned):
            current_year = AcademicYear.objects.order_by('-start_date').first()
            if not current_year:
                self.message_user(
                    request,
                    trans("No active academic year found. Please configure an academic year."),
                    level=messages.WARNING,
                )

        total_attempts = 0
        updated_attempts = 0
        for quiz in queryset:
            attempts = quiz.attempts.select_related('student')
            if not request.user.is_superuser:
                try:
                    teacher = Teacher.objects.get(user=request.user)
                    if current_year:
                        homeroom_classes = HomeroomTeacher.objects.filter(
                            teacher=teacher,
                            academic_year=current_year
                        ).values_list('school_class__id', flat=True)
                        attempts = attempts.filter(
                            student__enrollments__school_class__id__in=homeroom_classes,
                            student__enrollments__academic_year=current_year,
                            student__enrollments__status='ACTIVE'
                        ).distinct()
                    else:
                        attempts = attempts.none()
                except Teacher.DoesNotExist:
                    attempts = attempts.none()

            total_attempts += attempts.count()
            for attempt in attempts:
                old_score = attempt.score
                new_score = self.recalculate_quiz_attempt_score(attempt, request=request)
                if old_score != new_score:
                    updated_attempts += 1

        if total_attempts > 0:
            self.message_user(
                request,
                str(trans("បានគណនាពិន្ទុឡើងវិញសម្រាប់ %d QuizAttempt(s) និងធ្វើបច្ចុប្បន្នភាព %d នៅទូទាំងកម្រងសំណួរដែលបានជ្រើស។") % (total_attempts, updated_attempts)),
                level=messages.SUCCESS,
            )
        else:
            self.message_user(
                request,
                str(trans("គ្មាន QuizAttempt សម្រាប់គណនាឡើងវិញ។")),
                level=messages.WARNING,
            )

    recalculate_all_attempts.short_description = str(trans("គណនាពិន្ទុ All Attempt ម្ដងទៀត"))

    def recalculate_quiz_scores(self, request, object_id):
        """Handle manual recalculation of QuizAttempt scores for a specific quiz."""
        try:
            quiz = Quiz.objects.get(id=object_id)
            attempts = QuizAttempt.objects.filter(quiz=quiz)

            for attempt in attempts:
                # call the attempt-level recalculation
                self.recalculate_quiz_attempt_score(attempt)  # ONLY pass attempt

            messages.success(
                request,
                trans ("Successfully recalculated scores for %d QuizAttempt(s) for quiz: %s") % (len(attempts), quiz.title)
            )
        except Quiz.DoesNotExist:
            messages.error(request, trans ("Quiz not found."))
        return redirect(reverse('admin:quizzes_quiz_change', args=[object_id]))

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel),
                name="quizzes_import_excel",
            ),
            path(
                "download-template-quiz/",
                self.admin_site.admin_view(self.download_template),
                name="download_template_quiz",
            ),
            path(
                "<path:object_id>/recalculate-scores/",
                self.admin_site.admin_view(self.recalculate_quiz_scores),
                name="quizzes_recalculate_scores",
            ),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if not form.is_valid():
                messages.error(request,  trans ("មានបញ្ហាក្នុងទម្រង់បញ្ចូល: សូមពិនិត្យឯកសារដែលបានផ្ទុកឡើង"))
                return render(request, "admin/quizzes/import_quizzes.html", {"form": form})

            try:
                df = pd.read_excel(request.FILES["excel_file"]).fillna("")
                if df.empty:
                    messages.error(request,  trans ("ឯកសារ Excel ទទេ។ សូមផ្តល់ទិន្នន័យ"))
                    return render(request, "admin/quizzes/import_quizzes.html", {"form": form})

                column_map = {
                    "Category": "Category",
                    "Quiz Title": "Quiz Title",
                    "Description": "Description",
                    "Classes": "Classes",
                    "Time Limit": "Time Limit",
                    "Start Time": "Start Time",
                    "Status": "Status",
                    "Allow Check Answer": "Allow Check Answer",
                    "Allow See Score": "Allow See Score",
                    "Question Text": "Question Text",
                    "Question Type": "Question Type",
                    "Option Text": "Option Text",
                    "Is Correct": "Is Correct",
                    "Teacher Name": "Teacher Name",
                    "Points": "Points",
                    "Difficulty": "Difficulty",
                }
                df_columns = df.columns.str.strip()
                for col in df_columns:
                    for key in column_map:
                        if key.lower() in col.lower():
                            column_map[col] = column_map[key]

                required_columns = ["Quiz Title"]
                missing_columns = [col for col in required_columns if not any(col.lower() in c.lower() for c in df_columns)]
                if missing_columns:
                    messages.error(request,  trans ("ខ្វះជួរឈរ: %s. សូមប្រើទម្រង់ត្រឹមត្រូវ") % ", ".join(missing_columns))
                    return render(request, "admin/quizzes/import_quizzes.html", {"form": form})

                df = df.rename(columns=column_map)
                df = df[df[['Category', 'Quiz Title', 'Question Text']].notnull().any(axis=1)]

                current_quiz = None
                current_question = None
                answer_changed = False
                with transaction.atomic():
                    for index, row in df.iterrows():
                        try:
                            print(f"Row {index + 2}: Quiz Title={row.get('Quiz Title')}, Question Text={row.get('Question Text')}, "
                                  f"Status={row.get('Status')}, Is Correct={row.get('Is Correct')}, "
                                  f"Allow Check Answer={row.get('Allow Check Answer')}, Allow See Score={row.get('Allow See Score')}, "
                                  f"Teacher Name={row.get('Teacher Name')}")

                            if row.get("Quiz Title"):
                                category_name = row.get("Category") or "Imported"
                                category, _ = QuizCategory.objects.get_or_create(name=category_name)

                                teacher = None
                                if request.user.is_superuser:
                                    teacher_name = row.get("Teacher Name")
                                    if teacher_name:
                                        teacher = Teacher.objects.filter(given_name=teacher_name).first()
                                        if not teacher:
                                            messages.error(
                                                request,
                                                 trans ("គ្រូ '%s' មិនមាននៅក្នុងជួរទី %d") % (teacher_name, index + 2),
                                            )
                                            continue
                                else:
                                    teacher = Teacher.objects.filter(user=request.user).first()
                                    if not teacher:
                                        messages.error(
                                            request,
                                             trans ("គ្មានទម្រង់គ្រូសម្រាប់អ្នកប្រើប្រាស់នៅជួរទី %d") % (index + 2),
                                        )
                                        continue

                                class_names = row.get("Classes", "")
                                classes = []
                                if isinstance(class_names, str) and class_names.strip():
                                    class_names = [name.strip() for name in class_names.split(",")]
                                    classes = SchoolClass.objects.filter(name__in=class_names)
                                    if class_names and not classes:
                                        messages.warning(
                                            request,
                                             trans ("គ្មានថ្នាក់ត្រឹមត្រូវសម្រាប់ '%s' នៅជួរទី %d") % (row.get("Classes"), index + 2),
                                        )

                                time_limit = None
                                time_limit_str = row.get("Time Limit", "")
                                if time_limit_str:
                                    try:
                                        h, m, s = map(int, str(time_limit_str).split(":"))
                                        if h < 0 or m < 0 or s < 0 or m > 59 or s > 59:
                                            raise ValueError
                                        time_limit = timedelta(hours=h, minutes=m, seconds=s)
                                    except ValueError:
                                        messages.warning(
                                            request,
                                             trans ("ទម្រង់កំណត់ពេលវេលាមិនត្រឹមត្រូវនៅជួរទី %d: %s") % (index + 2, time_limit_str),
                                        )

                                start_time = None
                                start_time_str = row.get("Start Time", "")
                                if start_time_str:
                                    try:
                                        start_time = parse_datetime(str(start_time_str))
                                        if not start_time:
                                            raise ValueError("Invalid datetime format")
                                    except ValueError:
                                        messages.warning(
                                            request,
                                             trans ("ទម្រង់ពេលវេលាចាប់ផ្តើមមិនត្រឹមត្រូវនៅជួរទី %d: %s") % (index + 2, start_time_str),
                                        )

                                status = str(row.get("Status", "DRAFT")).upper()
                                status_choices = dict(Quiz.STATUS_CHOICES)
                                if status not in status_choices:
                                    messages.warning(
                                        request,
                                         trans ("ស្ថានភាពមិនត្រឹមត្រូវនៅជួរទី %d: %s. ប្រើ DRAFT") % (index + 2, status),
                                    )
                                    status = "DRAFT"

                                allow_check_answer = row.get("Allow Check Answer", False)
                                if isinstance(allow_check_answer, str):
                                    allow_check_answer = allow_check_answer.strip().lower() in ('true', '1', 'yes')
                                allow_see_score = row.get("Allow See Score", False)
                                if isinstance(allow_see_score, str):
                                    allow_see_score = allow_see_score.strip().lower() in ('true', '1', 'yes')

                                current_quiz, created = Quiz.objects.get_or_create(
                                    title=row["Quiz Title"],
                                    category=category,
                                    teacher=teacher,
                                    defaults={
                                        "description": row.get("Description", ""),
                                        "time_limit": time_limit,
                                        "start_time": start_time,
                                        "status": status,
                                        "allow_check_answer": allow_check_answer,
                                        "allow_see_score": allow_see_score,
                                    },
                                )
                                if not created:
                                    original_questions = {q.id: {'text': q.text, 'question_type': q.question_type} for q in current_quiz.questions.all()}
                                    original_options = {opt.id: {'is_correct': opt.is_correct, 'text': opt.text} for q in current_quiz.questions.all() for opt in q.options.all()}
                                if classes:
                                    current_quiz.classes.set(classes)
                                current_quiz.save()

                            if row.get("Question Text") and current_quiz:
                                points = row.get("Points", 1)
                                try:
                                    points = int(float(points)) if points else 1
                                except (ValueError, TypeError):
                                    points = 1
                                    messages.warning(
                                        request,
                                         trans ("ពិន្ទុមិនត្រឹមត្រូវនៅជួរទី %d: %s. ប្រើ 1") % (index + 2, row.get("Points")),
                                    )
                                question_type = row.get("Question Type", "MCQ_SINGLE")
                                if question_type not in dict(Question.QUESTION_TYPES):
                                    messages.warning(
                                        request,
                                         trans ("ប្រភេទសំណួរមិនត្រឹមត្រូវនៅជួរទី %d: %s. ប្រើ MCQ_SINGLE") % (index + 2, question_type),
                                    )
                                    question_type = "MCQ_SINGLE"
                                # Set difficulty to MEDIUM by default for imported questions
                                difficulty = row.get("Difficulty", "MEDIUM")
                                if difficulty not in dict(Question.DIFFICULTY_CHOICES):
                                    messages.warning(
                                        request,
                                        trans ("កម្រិតលំបាកមិនត្រឹមត្រូវនៅជួរទី %d: %s. ប្រើ MEDIUM") % (index + 2, difficulty),
                                    )
                                    difficulty = "MEDIUM"
                            
                                current_question, created = Question.objects.get_or_create(
                                    quiz=current_quiz,
                                    text=row["Question Text"],
                                    question_type=question_type,
                                    difficulty=difficulty,
                                    defaults={"points": points, "order": current_quiz.questions.count() + 1},
                                )
                                if not created and question_type == "SHORT":
                                    if original_questions.get(current_question.id, {}).get('text') != row["Question Text"]:
                                        answer_changed = True
                                        LogEntry.objects.log_action(
                                            user_id=request.user.id,
                                            content_type_id=ContentType.objects.get_for_model(Question).pk,
                                            object_id=current_question.pk,
                                            object_repr=str(current_question),
                                            action_flag=CHANGE,
                                            change_message= trans ("Updated SHORT question text to: %s") % row["Question Text"]
                                        )

                            if row.get("Option Text") and current_question:
                                is_correct_value = row.get("Is Correct", False)
                                if isinstance(is_correct_value, str):
                                    is_correct = is_correct_value.strip().lower() in ('true', '1', 'yes')
                                else:
                                    is_correct = bool(is_correct_value)
                                option, created = AnswerOption.objects.get_or_create(
                                    question=current_question,
                                    text=str(row["Option Text"]).strip(),
                                    defaults={"is_correct": is_correct},
                                )
                                if not created and original_options.get(option.id, {}).get('is_correct') != is_correct:
                                    answer_changed = True
                                    option.is_correct = is_correct
                                    option.save()
                                    LogEntry.objects.log_action(
                                        user_id=request.user.id,
                                        content_type_id=ContentType.objects.get_for_model(AnswerOption).pk,
                                        object_id=option.pk,
                                        object_repr=str(option),
                                        action_flag=CHANGE,
                                        change_message= trans ("Changed is_correct to %s for option: %s") % (is_correct, option.text)
                                    )
                                if not created and option.question.question_type == "SHORT" and original_options.get(option.id, {}).get('text') != option.text:
                                    answer_changed = True
                                    LogEntry.objects.log_action(
                                        user_id=request.user.id,
                                        content_type_id=ContentType.objects.get_for_model(AnswerOption).pk,
                                        object_id=option.pk,
                                        object_repr=str(option),
                                        action_flag=CHANGE,
                                        change_message= trans ("Updated SHORT answer text to: %s") % option.text
                                    )

                        except Exception as e:
                            messages.error(request,  trans ("មានបញ្ហានៅជួរទី %d: %s") % (index + 2, str(e)))
                            print(f"Error in row {index + 2}: {str(e)}")

                    if answer_changed:
                        attempts = QuizAttempt.objects.filter(quiz=current_quiz)
                        for attempt in attempts:
                            self.recalculate_quiz_attempt_score(attempt, request)
                        if attempts:
                            messages.info(request,  trans ("Recalculated scores for %d QuizAttempt(s) due to answer changes in import.") % len(attempts))

                    messages.success(request,  trans ("ការនាំចូល Excel សម្រេចដោយជោគជ័យ! បានដំណើរការជួរចំនួន %d") % len(df))
                    return redirect("..")
            except Exception as e:
                messages.error(request,  trans ("មានបញ្ហាអាន Excel: %s") % str(e))
                print(f"Excel read error: {str(e)}")
                return render(request, "admin/quizzes/import_quizzes.html", {"form": form})
        else:
            form = ExcelImportForm()
        return render(request, "admin/quizzes/import_quizzes.html", {"form": form})

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Quiz Import Template"

            headers = [
                "Category", "Teacher Name", "Quiz Title", "Description", "Classes",
                "Time Limit", "Start Time", "Status", "Allow Check Answer", "Allow See Score",
                "Question Text", "Question Type", "Option Text", "Is Correct", "Points"
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 15)

            sample_data = [
                ["Mathematics", "Mok Kong", "Basic Math Quiz", "Test math skills", "Class 16A,Class 16B", "00:30:00", "2025-09-18 09:00:00", "PUBLISH", False, False, "What is 2 + 2?", "MCQ_SINGLE", "4", True, 1],
                ["", "", "", "", "", "", "", "", "", "", "", "3", False, ""],
                ["", "", "", "", "", "", "", "", "", "What is 5 x 3?", "MCQ_SINGLE", "15", True, 2],
                ["", "", "", "", "", "", "", "", "", "", "", "18", False, ""],
                ["", "", "", "", "", "", "", "", "", "What is the sum of angles in a triangle?", "SHORT", "", "", 2],
                ["Science", "", "General Science Quiz", "Test science knowledge", "Class 16B", "00:45:00", "2025-09-19 10:00:00", "RELEASE", False, False, "Which are fruits? (Select all)", "MCQ_MULTI", "Apple", True, 2],
                ["", "", "", "", "", "", "", "", "", "", "", "Carrot", False, ""],
                ["", "", "", "", "", "", "", "", "", "", "", "Banana", True, ""],
                ["", "", "", "", "", "", "", "", "", "What is the chemical symbol for water?", "SHORT", "", "", 1],
                ["", "", "", "", "", "", "", "", "", "Which planets are gas giants? (Select all)", "MCQ_MULTI", "Earth", False, 2],
                ["", "", "", "", "", "", "", "", "", "", "", "Jupiter", True, ""],
                ["", "", "", "", "", "", "", "", "", "", "", "Saturn", True, ""],
                ["", "", "", "", "", "", "", "", "", "What is Mars primarily composed of?", "SHORT", "", "", 1],
                ["", "", "", "", "", "", "", "", "", "What is the speed of light?", "MCQ_SINGLE", "300,000 km/s", True, 1],
                ["", "", "", "", "", "", "", "", "", "", "", "250,000 km/s", False, ""],
            ]

            for row_num, row_data in enumerate(sample_data, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell_value = str(value) if value is not None else ""
                    ws.cell(row=row_num, column=col_num, value=cell_value)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return FileResponse(
                buffer,
                as_attachment=True,
                filename="quiz_import_template.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            import traceback
            print(f"Error in download_template: {e}\n{traceback.format_exc()}")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Error creating template. Check logs."
            ws['A2'] = str(e)
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return FileResponse(buffer, as_attachment=True, filename="error_template.xlsx")

# @admin.register(QuizAttempt)
class QuizAttemptAdmin(admin.ModelAdmin):
    list_display = ("student", "quiz", "score", "completed_at")
    list_filter = ("quiz", "completed_at")
    search_fields = ("student__username",)

    def student(self, obj):
        return obj.student
    student.short_description = "Student"

    def quiz(self, obj):
        return obj.quiz
    quiz.short_description = "Quiz"

# @admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ("text", "quiz", "question_type", "difficulty", "points")
    list_filter = ("quiz", "question_type", "difficulty")
    search_fields = ("text", "quiz__title")
    inlines = [AnswerOptionInline]

    def quiz(self, obj):
        return obj.quiz
    quiz.short_description = "Quiz"

# @admin.register(AnswerOption)
class AnswerOptionAdmin(admin.ModelAdmin):
    list_display = ("text", "question", "is_correct")
    list_filter = ("question__quiz", "is_correct")
    search_fields = ("text", "question__text")

    def question(self, obj):
        return obj.question
    question.short_description = "Question"


@admin.register(StudentResponse)
class StudentResponseAdmin(admin.ModelAdmin):
    list_display = ("student", "quiz", "question", "question_type")
    list_filter = ("attempt__quiz", "question__question_type", "attempt__student")

    def student(self, obj):
        return obj.student
    student.short_description = trans("Student")

    def quiz(self, obj):
        return obj.quiz
    quiz.short_description = trans("Quiz")

    def question_type(self, obj):
        return obj.question.question_type
    question_type.short_description =  trans ("Question Type")

    def selected_options_display(self, obj):
        return ", ".join(opt.text for opt in obj.selected_options.all()) or  trans ("None")
    selected_options_display.short_description =  trans ("Selected Options")