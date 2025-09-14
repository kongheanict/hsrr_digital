from django.shortcuts import render

# Create your views here.
# backend/teachers/views.py
import io
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from apps.teachers.models import Position

def download_teacher_template(request):
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers Template"

    # Column headers
    headers = [
        "TID",
        "Family name",
        "Given name",
        "Status",
        "ID card number",
        "Date of birth",
        "Email",
        "Gender",
        "Phone number",
        "Place of birth",
        "Specialized division",
        "Position",
        "Enrolled date",
    ]
    ws.append(headers)

    # -------------------
    # Styles
    # -------------------
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue
    header_font = Font(color="FFFFFF", bold=True)  # White bold
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Apply style to header row
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set row height for header
    ws.row_dimensions[1].height = 25

    # Adjust column widths (you can tweak values as needed)
    column_widths = [12, 20, 20, 15, 20, 15, 25, 12, 18, 25, 25, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width  # A=65 ASCII → chr(65)=A

    # -------------------
    # Data validation lists
    # -------------------
    # Status dropdown
    status_list = ["ACTIVE", "INACTIVE", "RETIRED", "MOVED"]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(status_list)}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("D2:D1000")  # Column D (Status)

    # Gender dropdown
    gender_list = ["Male", "Female", "Other"]
    dv_gender = DataValidation(type="list", formula1=f'"{",".join(gender_list)}"', allow_blank=True)
    ws.add_data_validation(dv_gender)
    dv_gender.add("H2:H1000")  # Column H (Gender)

    # Position dropdown from DB
    positions = list(Position.objects.values_list("name", flat=True))
    if positions:
        dv_position = DataValidation(type="list", formula1=f'"{",".join(positions)}"', allow_blank=True)
        ws.add_data_validation(dv_position)
        dv_position.add("L2:L1000")  # Column L (Position)

    # -------------------
    # Sample row
    # -------------------
    ws.append([
        "123", "Doe", "John", "ACTIVE", "123456789", "1997-09-22",
        "john@example.com", "Male", "0123456789", "City Name",
        "Math, Science", "Teacher", "2020-08-01"
    ])

    # Save workbook to in-memory file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename="teachers_template.xlsx")
